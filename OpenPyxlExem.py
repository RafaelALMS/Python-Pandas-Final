from IPython.display import display 
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment


combustiveis_df = pd.read_excel("ca202102_20230207120945.xlsx") #Coloca a bd em uma vareavel
c_mean = combustiveis_df['Valor de Venda'].groupby(by=combustiveis_df['Produto']).mean()

display(c_mean)
excel='por_litro.xlsx'
c_mean.to_excel(excel, 'Sumario')

#Vai abrir o Excel no openpyxl
wb = load_workbook(excel) # wb = Workbook

#Pegar a planilha certa... usando o Sheet Name (nome da planilha)
ws = wb['Sumário'] # Work Sheet -> planilha atual, ativa, de trabalho

#Vamos pintar o cabeçalho da tabela de "cinzinha"
cinzinha = PatternFill("solid", fgColor="CCCCCC")
coords = ['A1', 'B1']
for coord in coords:
  ws[coord].fill = cinzinha

#Salvar o Excel
wb.save(excel)

